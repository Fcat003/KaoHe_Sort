#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import os
import tempfile
import shutil


class ExcelSorterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("考核表 排序工具")
        self.root.geometry("550x300")
        self.root.resizable(False, False)

        self.setup_ui()
        self.center_window()

    def center_window(self):
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'+{x}+{y}')

    def setup_ui(self):
        # 创建主框架
        main_frame = ttk.Frame(self.root, padding=(20, 15))
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 创建标题
        title_label = ttk.Label(
            main_frame,
            text="考核表 排序工具",
            font=("Arial", 14, "bold")
        )
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 15))

        # 输入文件部分
        input_frame = ttk.Frame(main_frame)
        input_frame.grid(row=1, column=0, columnspan=3, sticky="ew", pady=5)

        ttk.Label(input_frame, text="源文件:").pack(side=tk.LEFT, padx=(0, 5))

        self.input_var = tk.StringVar()
        input_entry = ttk.Entry(input_frame, textvariable=self.input_var, width=45)
        input_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))

        input_btn = ttk.Button(input_frame, text="浏览...", command=self.choose_input)
        input_btn.pack(side=tk.LEFT)

        # 输出文件部分
        output_frame = ttk.Frame(main_frame)
        output_frame.grid(row=2, column=0, columnspan=3, sticky="ew", pady=5)

        ttk.Label(output_frame, text="保存为:").pack(side=tk.LEFT, padx=(0, 5))

        self.output_var = tk.StringVar()
        output_entry = ttk.Entry(output_frame, textvariable=self.output_var, width=45)
        output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))

        output_btn = ttk.Button(output_frame, text="浏览...", command=self.choose_output)
        output_btn.pack(side=tk.LEFT)

        # 工作表选择部分
        sheet_frame = ttk.Frame(main_frame)
        sheet_frame.grid(row=3, column=0, columnspan=3, pady=10, sticky="w")

        ttk.Label(sheet_frame, text="选择要排序的工作表:").pack(side=tk.LEFT, padx=(0, 10))

        self.sheet_var = tk.StringVar(value="非正职公务员")
        sheets = ["非正职公务员", "主要领导", "以上全部"]
        sheet_opt_frame = ttk.Frame(sheet_frame)
        sheet_opt_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)

        for sheet in sheets:
            rb = ttk.Radiobutton(
                sheet_opt_frame,
                text=sheet,
                variable=self.sheet_var,
                value=sheet
            )
            rb.pack(side=tk.LEFT, padx=5)

        # # 分组排序选项
        # self.group_var = tk.BooleanVar(value=True)
        # group_cb = ttk.Checkbutton(
        #     main_frame,
        #     text="按单位分组排序（仅适用于非正职公务员）",
        #     variable=self.group_var
        # )
        # group_cb.grid(row=4, column=0, columnspan=3, pady=(5, 10), sticky="w")

        # 进度条
        self.progress_var = tk.DoubleVar()
        progress_bar = ttk.Progressbar(
            main_frame,
            variable=self.progress_var,
            maximum=100,
            length=500,
            mode='determinate'
        )
        progress_bar.grid(row=5, column=0, columnspan=3, pady=(10, 5), sticky="ew")

        # 状态栏
        self.status_var = tk.StringVar(value="就绪")
        status_bar = ttk.Label(
            main_frame,
            textvariable=self.status_var,
            relief=tk.SUNKEN,
            anchor=tk.W,
            padding=(5, 2)
        )
        status_bar.grid(row=6, column=0, columnspan=3, sticky="ew", pady=(5, 0))

        # 操作按钮
        btn_frame = ttk.Frame(main_frame)
        btn_frame.grid(row=7, column=0, columnspan=3, pady=(15, 5))

        run_btn = ttk.Button(
            btn_frame,
            text="开始排序",
            command=self.run_sort,
            width=15
        )
        run_btn.pack(side=tk.LEFT, padx=10)

        exit_btn = ttk.Button(
            btn_frame,
            text="退出程序",
            command=self.root.destroy,
            width=15
        )
        exit_btn.pack(side=tk.LEFT, padx=10)

    def choose_input(self):
        path = filedialog.askopenfilename(
            title="选择源文件",
            filetypes=[("Excel 文件", "*.xlsx;*.xlsm;*.xltx;*.xltm")]
        )
        if path:
            self.input_var.set(path)
            if not self.output_var.get():
                dir_name, file_name = os.path.split(path)
                base_name, ext = os.path.splitext(file_name)
                new_name = f"{base_name}_排序后{ext}"
                self.output_var.set(os.path.join(dir_name, new_name))

    def choose_output(self):
        path = filedialog.asksaveasfilename(
            title="选择保存路径",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")]
        )
        if path:
            self.output_var.set(path)

    def run_sort(self):
        in_path = self.input_var.get()
        out_path = self.output_var.get()
        sheet = self.sheet_var.get()

        if not in_path or not out_path:
            messagebox.showwarning("提示", "请先选择输入文件和保存路径。")
            return

        try:
            self.status_var.set("正在处理，请稍候...")
            self.progress_var.set(0)
            self.root.update()

            if sheet == "以上全部":
                temp_path = tempfile.mktemp(suffix=".xlsx")
                shutil.copyfile(in_path, temp_path)

                self.status_var.set("正在排序非正职公务员...")
                self.root.update()
                sort_sheet(in_path, out_path, "非正职公务员", True)
                self.progress_var.set(50)

                self.status_var.set("正在排序主要领导...")
                self.root.update()
                # 直接从原文件读取分数，确保缓存完整
                sort_sheet(out_path, out_path, "主要领导", False)
                self.progress_var.set(100)

                os.remove(temp_path)
            else:
                group_by = True if sheet == "非正职公务员" else False
                self.status_var.set(f"正在排序{sheet}...")
                self.root.update()
                sort_sheet(in_path, out_path, sheet, group_by)
                self.progress_var.set(100)

            self.status_var.set("排序完成！")
            messagebox.showinfo("成功", f"'{sheet}'表排序并保存完成！")
            self.status_var.set("就绪")
            self.progress_var.set(0)
        except Exception as e:
            self.status_var.set("处理出错")
            self.progress_var.set(0)
            messagebox.showerror("错误", f"处理文件时出错：\n{str(e)}")
            self.status_var.set("就绪")


def sort_sheet(filepath, savepath, sheet_name, group_by_unit=False):
    # 读取 workbook，不依赖 data_only
    wb = load_workbook(filepath)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"工作簿中不存在工作表 '{sheet_name}'")
    ws = wb[sheet_name]

    # 表头行
    HDR_ROW = 6 if sheet_name != "主要领导" else 4
    COL_A = 1
    COL_B = column_index_from_string("B")
    COL_UNIT = column_index_from_string("C")
    # 计算得分使用 D/E/F/G
    COL_D = column_index_from_string("D")
    COL_E = column_index_from_string("E")
    COL_F = column_index_from_string("F")
    COL_G = column_index_from_string("G")

    COL_NOTE = column_index_from_string("I") if sheet_name != "主要领导" else None
    merged_ranges = ws.merged_cells.ranges

    def is_merged_non_anchor(r, c):
        for m in merged_ranges:
            if m.min_row <= r <= m.max_row and m.min_col <= c <= m.max_col:
                return not (r == m.min_row and c == m.min_col)
        return False

    # 找到最后一行
    last_data_row = HDR_ROW
    for r in range(ws.max_row, HDR_ROW, -1):
        if ws.cell(r, COL_G).value not in (None, ""):
            last_data_row = r
            break

    rows = []
    for rr in range(HDR_ROW + 1, last_data_row + 1):
        raw_vals = [ws.cell(rr, c).value for c in range(1, ws.max_column + 1)]
        # 处理合并单元格
        if group_by_unit:
            for col in (COL_B, COL_UNIT):
                if raw_vals[col - 1] in (None, ""):
                    for m in merged_ranges:
                        if m.min_col == col and m.min_row <= rr <= m.max_row:
                            raw_vals[col - 1] = ws.cell(m.min_row, m.min_col).value
                            break
        note = raw_vals[COL_NOTE - 1] if COL_NOTE else ""

        # **改动：在代码中直接计算分数，替代 data_only 读取**
        try:
            if sheet_name == "主要领导":
                score = (raw_vals[COL_D-1] or 0) * 0.35 + (raw_vals[COL_E-1] or 0) * 0.25 + (raw_vals[COL_F-1] or 0) * 0.20 + (raw_vals[COL_G-1] or 0) * 0.20
            else:
                score = (raw_vals[COL_E-1] or 0) * 0.35 + (raw_vals[COL_F-1] or 0) * 0.30 + (raw_vals[COL_G-1] or 0) * 0.35
        except Exception:
            score = 0.0

        rows.append((rr, raw_vals, note, score))

    # 排序逻辑保持不变
    if group_by_unit:
        sorted_rows = []
        groups, start, last_unit = [], 0, None
        for i, (_, vals, _, _) in enumerate(rows):
            unit = vals[COL_UNIT - 1]
            if unit and unit != last_unit and last_unit is not None:
                groups.append((start, i))
                start = i
            last_unit = unit or last_unit
        groups.append((start, len(rows)))
        for s, e in groups:
            grp = rows[s:e]
            grp.sort(key=lambda x: (bool(x[2]), -x[3]))
            sorted_rows.extend(grp)
    else:
        sorted_rows = sorted(rows, key=lambda x: -x[3])

    # 写回数据
    for idx, (_orig, vals, _, _) in enumerate(sorted_rows, start=HDR_ROW + 1):
        for c, v in enumerate(vals, start=1):
            if not is_merged_non_anchor(idx, c):
                ws.cell(idx, c).value = v

    # 重写序号与公式
    for i, (_orig, _vals, _note, _score) in enumerate(sorted_rows, start=1):
        r = HDR_ROW + i
        if not is_merged_non_anchor(r, COL_A):
            ws.cell(r, COL_A).value = i
        if sheet_name == "主要领导":
            ws.cell(r, column_index_from_string("H")).value = f"=D{r}*35%+E{r}*25%+F{r}*20%+G{r}*20%"
        else:
            ws.cell(r, column_index_from_string("H")).value = f"=SUM(E{r}*35%+F{r}*30%+G{r}*35%)"

    wb.save(savepath)
    wb.close()


if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelSorterApp(root)
    root.mainloop()
